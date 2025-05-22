from openpyxl import load_workbook


def get_excel_data(path: str) -> list[tuple]:
    wb = load_workbook(path)
    data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            data.append(row)

    return data
