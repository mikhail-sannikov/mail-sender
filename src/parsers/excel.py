from openpyxl import load_workbook


def get_receivers() -> list[tuple]:
    wb = load_workbook('../templates/receivers.xlsx')
    emails = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            emails.append(row)

    return emails
